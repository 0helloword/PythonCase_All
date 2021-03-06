#coding:utf-8
import random
import os
import time
import datetime
from os.path import join
from xlutils.copy import copy
import sys from openpyxl 
import Workbook, load_workbook from contextlib import closing
import openpyxl
import xlutils
import win32com.client as win32 from openpyxl.styles 
import colors from openpyxl.styles 
import Font, Color from openpyxl.styles 
import colors from openpyxl 
import Workbook
import subprocess

reload(sys)
sys.setdefaultencoding('utf-8')

class recordLoanData(object):
    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    ROBOT_LIBRARY_VERSION = 0.1


    def __init__(self):
        pass

    ## 创建Excel
    def createExcel(self, file_name):
        with closing(Workbook()) as wb:
            wb.save(file_name)
    ## 获取Excel行数
    def getExcelRowCount(self, file_name, sheet_name):
        with closing(load_workbook(filename=file_name)) as wb:
            rows = wb.get_sheet_by_name(name = str(sheet_name)).max_row
            return rows
    ## 获取Excel列数
    def getExcelColumnCount(self, file_name, sheet_name):
        with closing(load_workbook(filename=file_name)) as wb:
            columns = wb.get_sheet_by_name(name = str(sheet_name)).max_column
            return columns
    ## 按照行列读取Excel某一单元格的值
    def readExcel(self, file_name, sheet_name, cell_row, cell_column):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.active
            ws.title = str(sheet_name)
            cellValue = str(ws.cell(row=int(cell_row),column=int(cell_column)).value)
            return cellValue.decode('utf-8')

    ## 按照行列向某一单元格写入值
    def addExcel(self, file_name, sheet_name, cell_row, cell_column, value):
        with closing(load_workbook(filename=file_name)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            ft = Font(name="微软雅黑 Light", size=9)
            ws.cell(row=int(cell_row), column=int(cell_column), value=value).font = ft
            wb.save(file_name)

    ## 读取从某一行某一列开始的n个值（n=cell_column）
    def listReadExcel(self, file_name, sheet_name, cell_row, cell_column, start_col):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            cellValueList = []
            for i in range(cell_column):
                cellValue = ws.cell(row=int(cell_row),column=int(start_col)+i).value
                cellValueList.append(cellValue)
            return cellValueList

    ## 从某一行的某一列开始写入n个值（*value）
    def listAddExcel(self, file_name, sheet_name, cell_row, cell_column, *value):
        with closing(load_workbook(filename=file_name)) as wb:
            sheets = wb.get_sheet_names()
            if sheet_name in sheets:
                ws = wb.get_sheet_by_name(str(sheet_name))
            else:
                ws = wb.create_sheet(str(sheet_name))
            ft1 = Font(name="微软雅黑", color=colors.RED, size=12)
            ft2 = Font(name="微软雅黑", size=12, bold=True)
            ft3 = Font(name="微软雅黑 Light", size=9)
            for i in range(len(value)):
                if value[i] == 'FAIL':
                    ws.cell(row=int(cell_row), column=int(cell_column)+i, value=value[i]).font = ft1
                elif value[i] == 'PASS':
                    ws.cell(row=int(cell_row), column=int(cell_column) + i, value=value[i]).font = ft2
                else:
                    ws.cell(row=int(cell_row), column=int(cell_column) + i, value=value[i]).font = ft3
            wb.save(file_name)

    ##刷新Excel文件,用于让Excel自动计算其中的公式
    def refreshExcel(self, file_name):
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        workbook = excel.Workbooks.Open(os.path.join(os.getcwd(), file_name))
        workbook.Save()
        excel.Application.Quit()

    #按照单元格名称写入Excel（如A5）
    def cellAddExcel(self, file_name, sheet_name, cell, value):
        with closing(load_workbook(filename=file_name)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            ft = Font(name="微软雅黑 Light", size=9, color=colors.GREEN)
            ws[str(cell)] = value
            ws[str(cell)].font = ft
            wb.save(file_name)


    #按照列名读取一列固定位置开始和结束的一段值(如 E 5 10)
    def cellListReadExcel(self, file_name, sheet_name, col_name, start, end):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            cellValueList = []
            for i in range(int(start),int(end)):
                cellname = str(col_name) + str(i)
                cellValue = ws[str(cellname)].value
                if cellValue == None:
                    cellValue=''
                cellValueList.append(cellValue)
            return cellValueList

    ##拷贝Excel中内容 把Excel中一个sheet页内容复制到另一个Excel
    def excelCopy(self, fromfile_name, tofile_name, fromsheet_name, tosheet_name):
        wb1 = load_workbook(str(fromfile_name))
        wb2 = load_workbook(str(tofile_name))
        ft = Font(name="微软雅黑", size=10)
        sheets = wb2.get_sheet_names()
        if tosheet_name in sheets:
            wb2.get_sheet_by_name(str(tosheet_name))
        else:
            wb2.create_sheet(str(tosheet_name))
        ws1 = wb1[str(fromsheet_name)]
        ws2 = wb2[str(tosheet_name)]
        for i, row in enumerate(ws1.iter_rows()):
            for j, cell in enumerate(row):
                ws2.cell(row=i + 1, column=j + 1, value=cell.value).font = ft
        wb2.save(tofile_name)

    #按行读取Excel中内容,行数从0开始 如第1行 col_count=0
    def rowsReadExcel(self, file_name, sheet_name, col_count):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            rowValueList = []
            num = int(col_count)
            for cell in list(ws.rows)[num]:
                values = cell.value
                rowValueList.append(values)
            return rowValueList

    # 按行读取Excel中非空内容,遇到空值就结束,即使后面有值,行数从0开始 如第1行 col_count=0
    # 可以自己修改，将break修改为continue,读取一行所有非空的值
    def rowsReadExcel2(self, file_name, sheet_name, col_count):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            rowValueList = []
            num = int(col_count)
            for cell in list(ws.rows)[num]:
                if cell.value != None:
                    values = cell.value
                else:
                    break
                rowValueList.append(values)
            return rowValueList

    #按列读取Excel中内容
    def colsReadExcel(self, file_name, sheet_name, row_count):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            colValueList = []
            num = int(row_count)
            for cell in list(ws.columns)[num]:
                values = cell.value
                colValueList.append(values)
            return colValueList

    #按列读取Excel中连续非空内容
    def colsReadExcel2(self, file_name, sheet_name, row_count):
        with closing(load_workbook(filename=file_name, data_only=True)) as wb:
            ws = wb.get_sheet_by_name(str(sheet_name))
            colValueList = []
            num = int(row_count)
            for cell in list(ws.columns)[num]:
                if cell.value != None:
                    values = cell.value
                else:
                    break
                colValueList.append(values)
            return colValueList

    #将测试结果写入到Excel
    def resultAddExcel(self, file_name, sheet_name, cell_row, cell_column, count, *value):
        with closing(load_workbook(filename=file_name)) as wb:
            sheets = wb.get_sheet_names()
            if sheet_name in sheets:
                ws = wb.get_sheet_by_name(str(sheet_name))
            else:
                ws = wb.create_sheet(str(sheet_name))
            ft1 = Font(name="微软雅黑", color=colors.RED, size=12)
            ft2 = Font(name="微软雅黑", size=12, bold=True)
            ft3 = Font(name="微软雅黑 Light", size=9)
            if value[int(count)] == 'FAIL':
                ws.cell(row=int(cell_row), column=int(cell_column), value=value[int(count)]).font = ft1
            elif value[int(count)] == 'PASS':
                ws.cell(row=int(cell_row), column=int(cell_column), value=value[int(count)]).font = ft2
            else:
                ws.cell(row=int(cell_row), column=int(cell_column), value=value[int(count)]).font = ft3
            wb.save(file_name)

    #将两个列表转换成字典
    def listToDict(self, list1, list2):
        L1 = list1
        L2 = list2
        return dict(zip(L1,L2))

    #获取列表元素个数
    def listLen(self,list):
        count = len(list)
        return count

    #通过字典的值获取对应Key
    def get_keys(self, d, val):
        return [k for k, v in d.items() if v == val]

    #将两个字典变量合并成一个新的字典
    def newDict(self, dict1, dict2):
        def get_keys(d, val):
            return [k for k, v in d.items() if v == val]
        newdict = {}
        for key in dict1.keys():
            for value in dict2.values():
                if key == value:
                    ls = dict1[key]
                    vkey = ls
                    kvalue = get_keys(dict2, value)
                    newdict[kvalue[0]] = str(vkey)
        return newdict

    #生成区间数字列表
    def listRange(self, start, end):
        listID = range(int(start), int(end))
        return listID

    #将字典变量内容写入到Excel
    def dictAddTitleExcel(self, file_name, sheet_name, cell_row, dict):
        with closing(load_workbook(filename=file_name)) as wb:
            sheets = wb.get_sheet_names()
            if sheet_name in sheets:
                ws = wb.get_sheet_by_name(str(sheet_name))
            else:
                ws = wb.create_sheet(str(sheet_name))
            ft = Font(name="微软雅黑 Light", size=10, bold=False)
            for key in dict.keys():
                val = dict[key]
                ws.cell(row=int(cell_row), column=int(key), value=str(val)).font = ft
            wb.save(file_name)

    #新建文件夹
    def mkDir(self,path):
        a = os.path.exists(path)
        if a:
            pass
        else:
            os.makedirs(path)